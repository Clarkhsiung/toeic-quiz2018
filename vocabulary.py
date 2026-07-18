import os
import time
import re
from faster_whisper import WhisperModel

# 引入免費且穩定的翻譯庫
try:
    from deep_translator import GoogleTranslator
    translator = GoogleTranslator(source='en', target='zh-TW')
except ImportError:
    translator = None

# 1. 初始化模型：啟用 GPU (CUDA) 加速
print("現在開始進行中英雙語轉字幕與單字本生成，正在載入 large-v3 模型...")
model = WhisperModel("large-v3", device="cuda", compute_type="float16")

# 2. 自動搜尋音檔
audio_extensions = (".mp3", ".wav", ".m4a", ".flac")
audio_files = []

for root, dirs, files in os.walk('.'):
    for file in files:
        if file.lower().endswith(audio_extensions):
            full_path = os.path.join(root, file)
            audio_files.append(full_path)

if not audio_files:
    print("找不到任何音檔，請確認程式碼是否放對位置。")
    exit()

print(f"共找到 {len(audio_files)} 個音檔，開始處理...\n")

# 3. 時間戳記工具
def format_srt_timestamp(seconds):
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    secs = int(seconds % 60)
    millis = int((seconds % 1) * 1000)
    return f"{hours:02d}:{minutes:02d}:{secs:02d},{millis:03d}"

def format_lrc_timestamp(seconds):
    minutes = int(seconds // 60)
    secs = int(seconds % 60)
    millis = int((seconds % 1) * 100)
    return f"[{minutes:02d}:{secs:02d}.{millis:02d}]"

def extract_english_words(segments):
    """從語音辨識片段中提取不重複的英文單字並排序"""
    words_set = set()
    for seg in segments:
        words = re.findall(r'[a-zA-Z]+', seg.text)
        for w in words:
            if len(w) > 1:
                words_set.add(w.lower())
    return sorted(list(words_set))

def generate_vocabulary_excel(excel_path, words_list, audio_name):
    """生成包含自動中文翻譯的精美 Excel 檔"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "多益單字本"
    ws.views.sheetView[0].showGridLines = True
    
    navy_dark = "1B365D"
    ice_blue = "F0F4F8"
    border_color = "D9D9D9"
    
    title_font = Font(name="Microsoft JhengHei", size=15, bold=True, color="1B365D")
    header_font = Font(name="Microsoft JhengHei", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Microsoft JhengHei", size=11, color="333333")
    
    ws['A1'] = "TOEIC 多益雙語對照單字本"
    ws['A1'].font = title_font
    ws['A2'] = f"音檔來源: {audio_name}"
    ws['A2'].font = Font(name="Microsoft JhengHei", size=10, italic=True, color="666666")
    
    headers = ["序號", "英文單字 (Vocabulary)", "中文釋義 (AI 自動翻譯)", "學習熟練度"]
    ws.append([])
    ws.append(headers)
    
    header_fill = PatternFill(start_color=navy_dark, end_color=navy_dark, fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color=border_color), right=Side(style='thin', color=border_color),
        top=Side(style='thin', color=border_color), bottom=Side(style='thin', color=border_color)
    )
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # 填入單字並進行線上翻譯
    for idx, word in enumerate(words_list, start=1):
        row_idx = 4 + idx
        
        # 線上查詢中文意義
        chinese_meaning = ""
        if translator:
            try:
                chinese_meaning = translator.translate(word)
                time.sleep(0.1) # 溫和查詢防止被 Google 封鎖
            except Exception:
                chinese_meaning = "（翻譯失敗，請手動補齊）"
        
        ws.cell(row=row_idx, column=1, value=idx)
        ws.cell(row=row_idx, column=2, value=word)
        ws.cell(row=row_idx, column=3, value=chinese_meaning)
        ws.cell(row=row_idx, column=4, value="未學習")
        
        row_fill = PatternFill(start_color=ice_blue, end_color=ice_blue, fill_type="solid") if idx % 2 == 0 else PatternFill(fill_type=None)
        
        for col_idx in range(1, 5):
            c = ws.cell(row=row_idx, column=col_idx)
            c.font = body_font
            c.border = thin_border
            if row_fill.fill_type:
                c.fill = row_fill
            
            if col_idx in [1, 4]:
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
                
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    
    wb.save(excel_path)

# 4. 批次處理
start_total_time = time.time()
success_count = 0

for idx, audio_path in enumerate(audio_files, start=1):
    file_base = os.path.splitext(audio_path)[0]
    audio_name = os.path.basename(audio_path)
    
    srt_path = f"{file_base}.srt"
    lrc_path = f"{file_base}.lrc"
    excel_path = f"{file_base}_vocabulary.xlsx"
    
    print(f"[{idx}/{len(audio_files)}] 正在處理: {audio_path}")
    start_file_time = time.time()
    
    try:
        segments, info = model.transcribe(
            audio_path, beam_size=5, language="en",
            condition_on_previous_text=False, no_speech_threshold=0.6, compression_ratio_threshold=2.4
        )
        segments = list(segments)
        
        # 輸出 SRT (電腦播放用)
        with open(srt_path, "w", encoding="utf-8") as srt_file:
            for seg_idx, segment in enumerate(segments, start=1):
                start_time = format_srt_timestamp(segment.start)
                end_time = format_srt_timestamp(segment.end)
                text = segment.text.strip()
                srt_file.write(f"{seg_idx}\n{start_time} --> {end_time}\n{text}\n\n")
                
        # 輸出 LRC (Android 手機用)
        with open(lrc_path, "w", encoding="utf-8") as lrc_file:
            lrc_file.write(f"[ti:{audio_name}]\n")
            for segment in segments:
                start_time_lrc = format_lrc_timestamp(segment.start)
                text = segment.text.strip()
                lrc_file.write(f"{start_time_lrc}{text}\n")
                
        # 提取單字、自動翻譯並寫入帶有中文意義的 Excel
        extracted_words = extract_english_words(segments)
        generate_vocabulary_excel(excel_path, extracted_words, audio_name)
                
        file_elapsed = time.time() - start_file_time
        print(f"   => 完成！已補齊中英文對照 Excel。耗時: {file_elapsed:.1f} 秒\n")
        success_count += 1
            
    except Exception as e:
        print(f"   => 錯誤訊息: {e}\n")
        continue

total_elapsed = time.time() - start_total_time
print(f"【全部總字幕與雙語單字本生成完畢】")
print(f"成功處理: {success_count}/{len(audio_files)} 個檔案，總花費：{total_elapsed/60:.1f} 分鐘！")