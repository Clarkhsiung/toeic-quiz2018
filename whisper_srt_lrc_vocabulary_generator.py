import os
import time
import re
from faster_whisper import WhisperModel

# 1. 初始化模型：啟用 GPU (CUDA) 加速
print("現在開始進行轉字幕功能，正在載入 large-v3 模型至 RTX 3060 Ti GPU...")
model = WhisperModel("large-v3", device="cuda", compute_type="float16")

# 2. 自動搜尋當前目錄及其所有子資料夾內的所有英文音檔
audio_extensions = (".mp3", ".wav", ".m4a", ".flac")
audio_files = []

for root, dirs, files in os.walk('.'):
    for file in files:
        if file.lower().endswith(audio_extensions):
            full_path = os.path.join(root, file)
            audio_files.append(full_path)

if not audio_files:
    print("找不到任何音檔，請確認程式碼是否放對位置。")
    exit()

print(f"共找到 {len(audio_files)} 個英文音檔，開始利用 GPU 進行抗幻聽辨識...\n")

# 3. 時間戳記格式化工具
def format_srt_timestamp(seconds):
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    secs = int(seconds % 60)
    millis = int((seconds % 1) * 1000)
    return f"{hours:02d}:{minutes:02d}:{secs:02d},{millis:03d}"

def format_lrc_timestamp(seconds):
    minutes = int(seconds // 60)
    secs = int(seconds % 60)
    millis = int((seconds % 1) * 100) # LRC uses 2 digits for centiseconds
    return f"[{minutes:02d}:{secs:02d}.{millis:02d}]"

def extract_english_words(segments):
    """從所有片段中提取不重複的英文單字並排序"""
    words_set = set()
    for seg in segments:
        # 使用正規表達式過濾出純英文單字，並轉為小寫以去重
        words = re.findall(r'[a-zA-Z]+', seg.text)
        for w in words:
            if len(w) > 1: # 過濾掉單個字母如 a, i (可依需求調整)
                words_set.add(w.lower())
    return sorted(list(words_set))

def generate_vocabulary_excel(excel_path, words_list, audio_name):
    """生成精美的英文單字複習 Excel 檔"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "單字本"
    
    # 啟用格線
    ws.views.sheetView[0].showGridLines = True
    
    # 標題與風格
    navy_dark = "1B365D"
    ice_blue = "F0F4F8"
    border_color = "D9D9D9"
    
    title_font = Font(name="Microsoft JhengHei", size=16, bold=True, color="1B365D")
    header_font = Font(name="Microsoft JhengHei", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Microsoft JhengHei", size=11, color="333333")
    
    # 寫入標題資訊
    ws['A1'] = "語音辨識英文單字本"
    ws['A1'].font = title_font
    ws['A2'] = f"來源檔案: {audio_name}"
    ws['A2'].font = Font(name="Microsoft JhengHei", size=10, italic=True, color="666666")
    
    # 標頭行
    headers = ["序號", "英文單字 (Vocabulary)", "中文釋義 (請自行填寫)", "學習熟練度"]
    ws.append([]) # 空行 A3
    ws.append(headers) # A4
    
    header_fill = PatternFill(start_color=navy_dark, end_color=navy_dark, fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color=border_color),
        right=Side(style='thin', color=border_color),
        top=Side(style='thin', color=border_color),
        bottom=Side(style='thin', color=border_color)
    )
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # 寫入資料
    for idx, word in enumerate(words_list, start=1):
        row_idx = 4 + idx
        ws.cell(row=row_idx, column=1, value=idx)
        ws.cell(row=row_idx, column=2, value=word)
        ws.cell(row=row_idx, column=3, value="") # 留白供使用者填寫
        ws.cell(row=row_idx, column=4, value="未學習") # 預設狀態
        
        # 樣式與斑馬紋
        row_fill = PatternFill(start_color=ice_blue, end_color=ice_blue, fill_type="solid") if idx % 2 == 0 else PatternFill(fill_type=None)
        
        for col_idx in range(1, 5):
            c = ws.cell(row=row_idx, column=col_idx)
            c.font = body_font
            c.border = thin_border
            if row_fill.fill_type:
                c.fill = row_fill
            
            if col_idx in [1, 4]:
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
                
    # 自動調整欄寬
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 4: continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 5, 15)
        
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    
    wb.save(excel_path)

# 4. 迴圈批次處理
start_total_time = time.time()
success_count = 0

for idx, audio_path in enumerate(audio_files, start=1):
    file_base = os.path.splitext(audio_path)[0]
    audio_name = os.path.basename(audio_path)
    srt_path = f"{file_base}.srt"
    lrc_path = f"{file_base}.lrc"
    excel_path = f"{file_base}_vocabulary.xlsx"
    
    print(f"[{idx}/{len(audio_files)}] 正在處理: {audio_path}")
    start_file_time = time.time()
    
    try:
        segments, info = model.transcribe(
            audio_path, 
            beam_size=5, 
            language="en",
            condition_on_previous_text=False,
            no_speech_threshold=0.6,
            compression_ratio_threshold=2.4
        )
        
        segments = list(segments)
        
        # 輸出一：用於電腦播放的 SRT 檔
        with open(srt_path, "w", encoding="utf-8") as srt_file:
            for seg_idx, segment in enumerate(segments, start=1):
                start_time = format_srt_timestamp(segment.start)
                end_time = format_srt_timestamp(segment.end)
                text = segment.text.strip()
                # 濃縮為單行字串傳送，徹底根絕跨行引號斷裂問題
                srt_file.write(f"{seg_idx}\n{start_time} --> {end_time}\n{text}\n\n")
                
        # 寫入 LRC (適合 Android 手機播放器)
        with open(lrc_path, "w", encoding="utf-8") as lrc_file:
            # 加入音樂資訊標頭
            lrc_file.write(f"[ti:{audio_name}]")
            for segment in segments:
                start_time_lrc = format_lrc_timestamp(segment.start)
                text = segment.text.strip()
                lrc_file.write(f"{start_time_lrc}{text}")
                
        # 提取英文單字並生成精美 Excel 檔
        all_words = extract_english_words(segments)
        generate_vocabulary_excel(excel_path, all_words, audio_name)
                
        file_elapsed = time.time() - start_file_time
        print(f"   => 完成！已生成 SRT、LRC 與單字 Excel 檔。耗時: {file_elapsed:.1f} 秒")
        success_count += 1
            
    except Exception as e:
        print(f"   => 警告！無法讀取該檔案。錯誤訊息: {e}")
        continue

total_elapsed = time.time() - start_total_time
print(f"【全部總字幕與單字本生成完畢】")
print(f"成功處理: {success_count}/{len(audio_files)} 個檔案")
print(f"總共花費：{total_elapsed/60:.1f} 分鐘！")
